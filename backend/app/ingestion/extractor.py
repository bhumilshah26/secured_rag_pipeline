"""Document extraction into titled SECTIONS.

Routes by file type and returns a list of `Section(title, text)` so indexing can be
section-aware. Supports:
  - PDF  : text via pypdf; OCR fallback (poppler -> images -> tesseract) for scanned pages
  - DOCX : paragraphs grouped by Heading styles
  - XLSX : one section per worksheet
  - images (png/jpg/tiff): OCR via tesseract
  - txt / md : markdown headings split into sections

OCR and Office parsers degrade gracefully: if a tool/lib is unavailable the extractor
falls back to whatever text it could get rather than crashing ingestion.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass

from app.config import settings


@dataclass
class Section:
    title: str
    text: str


# ---- helpers ----------------------------------------------------------------
def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _looks_like_heading(line: str) -> bool:
    s = line.strip()
    if not (0 < len(s) <= 80):
        return False
    if re.match(r"^\d+(\.\d+)*\.?\s+\S", s):          # 1, 1.2, 2.3.1 ...
        return True
    if re.match(r"^#{1,6}\s+\S", s):                   # markdown heading
        return True
    if s.isupper() and len(s.split()) <= 12:           # ALL CAPS heading
        return True
    return False


def _split_into_sections(text: str, default_title: str) -> list[Section]:
    """Heuristically split flat text into sections on heading-like lines."""
    sections: list[Section] = []
    title = default_title
    buf: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        if _looks_like_heading(line):
            if buf:
                sections.append(Section(title=title, text="\n".join(buf).strip()))
                buf = []
            title = re.sub(r"^#{1,6}\s+", "", line).strip()
        elif line.strip():
            buf.append(line)
    if buf:
        sections.append(Section(title=title, text="\n".join(buf).strip()))
    return [s for s in sections if s.text] or [Section(default_title, text.strip())]


# ---- OCR --------------------------------------------------------------------
def _configure_tesseract():
    import pytesseract

    if settings.tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = settings.tesseract_cmd
    return pytesseract


def _ocr_image_bytes(data: bytes) -> str:
    try:
        from PIL import Image

        pyt = _configure_tesseract()
        return pyt.image_to_string(Image.open(io.BytesIO(data)))
    except Exception:
        return ""


def _ocr_pdf(data: bytes) -> str:
    try:
        from pdf2image import convert_from_bytes

        pyt = _configure_tesseract()
        kwargs = {"poppler_path": settings.poppler_path} if settings.poppler_path else {}
        images = convert_from_bytes(data, **kwargs)
        return "\n".join(pyt.image_to_string(img) for img in images)
    except Exception:
        return ""


# ---- per-type extractors ----------------------------------------------------
def _extract_pdf(data: bytes) -> list[Section]:
    text = ""
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        pages = [(p.extract_text() or "") for p in reader.pages]
        text = "\n".join(pages).strip()
    except Exception:
        text = ""
    # Scanned PDF (little/no embedded text) -> OCR
    if settings.ocr_enabled and len(text) < 40:
        ocr = _ocr_pdf(data)
        if ocr.strip():
            text = ocr
    return _split_into_sections(text, "Document")


def _extract_docx(data: bytes) -> list[Section]:
    try:
        import docx
    except Exception:
        return _split_into_sections(data.decode("utf-8", "ignore"), "Document")

    document = docx.Document(io.BytesIO(data))
    sections: list[Section] = []
    title = "Document"
    buf: list[str] = []
    for para in document.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "").lower() if para.style else ""
        if style.startswith("heading") or style == "title":
            if buf:
                sections.append(Section(title, "\n".join(buf).strip()))
                buf = []
            title = text
        else:
            buf.append(text)
    if buf:
        sections.append(Section(title, "\n".join(buf).strip()))
    return [s for s in sections if s.text] or [Section("Document", "")]


def _extract_xlsx(data: bytes) -> list[Section]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [Section("Spreadsheet", "")]

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sections: list[Section] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            sections.append(Section(title=ws.title, text="\n".join(rows)))
    return sections or [Section("Spreadsheet", "")]


def _extract_plaintext(data: bytes) -> list[Section]:
    text = data.decode("utf-8", "ignore")
    lines = [ln.rstrip() for ln in text.replace("\r\n", "\n").split("\n")]
    return _split_into_sections("\n".join(lines).strip(), "Document")


# ---- public API -------------------------------------------------------------
_IMAGE_EXTS = {"png", "jpg", "jpeg", "tiff", "tif", "bmp", "gif"}


def extract_sections(*, raw: bytes | str, mime_type: str = "", filename: str = "") -> list[Section]:
    if isinstance(raw, str):
        return _split_into_sections(raw, filename or "Document")

    ext = _ext(filename)
    mime = (mime_type or "").lower()

    if ext == "pdf" or "pdf" in mime:
        return _extract_pdf(raw)
    if ext in {"docx"} or "wordprocessingml" in mime:
        return _extract_docx(raw)
    if ext in {"xlsx", "xlsm"} or "spreadsheetml" in mime:
        return _extract_xlsx(raw)
    if ext in _IMAGE_EXTS or mime.startswith("image/"):
        return [Section(filename or "Image", _ocr_image_bytes(raw))] if settings.ocr_enabled else []
    return _extract_plaintext(raw)


def extract_text(*, raw: bytes | str, mime_type: str = "text/plain") -> str:
    """Backwards-compatible flat-text helper (sections joined)."""
    sections = extract_sections(raw=raw, mime_type=mime_type)
    return "\n\n".join(f"{s.title}\n{s.text}" for s in sections).strip()
